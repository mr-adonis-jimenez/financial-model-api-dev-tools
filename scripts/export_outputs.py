"""
export_outputs.py
=================
Exports the metrics and simulated-inputs CSVs to a tidy JSON bundle and
an Excel-ready format (separate sheets) suitable for import into the
financial model workbook or for sharing with stakeholders.

Outputs
-------
- data/outputs/financial_summary.json  — full metrics as JSON
- data/outputs/export_<YYYY-MM-DD>.xlsx — Excel workbook with:
    • Sheet "Metrics"          — monthly P&L and unit economics
    • Sheet "Simulated Inputs" — per-tier monthly detail
    • Sheet "Assumptions"      — snapshot of assumptions.json

Usage
-----
    python scripts/export_outputs.py \
        [--metrics  data/outputs/metrics.csv] \
        [--inputs   data/outputs/simulated_inputs.csv] \
        [--out-dir  data/outputs/]
"""

from __future__ import annotations

import argparse
import csv
import json
from datetime import date
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
ASSUMPTIONS_PATH = ROOT / "data" / "inputs" / "assumptions.json"
DEFAULT_METRICS = ROOT / "data" / "outputs" / "metrics.csv"
DEFAULT_INPUTS_CSV = ROOT / "data" / "outputs" / "simulated_inputs.csv"
DEFAULT_OUT_DIR = ROOT / "data" / "outputs"


def load_assumptions(path: Path) -> dict[str, Any]:
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def read_csv(path: Path) -> list[dict[str, Any]]:
    with open(path, newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def write_json(data: Any, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2)


def write_excel(
    metrics: list[dict[str, Any]],
    inputs: list[dict[str, Any]],
    assumptions: dict[str, Any],
    out_dir: Path,
) -> Path:
    """Write a multi-sheet Excel workbook. Requires openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel export. Install it with: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()
    HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    def _write_sheet(ws: Any, rows: list[dict[str, Any]]) -> None:
        if not rows:
            return
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(list(row.values()))
        # Auto-fit column widths
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(header)), 10)
            ws.column_dimensions[col_letter].width = max_len + 2

    # Sheet 1 — Metrics
    ws_metrics = wb.active
    ws_metrics.title = "Metrics"
    _write_sheet(ws_metrics, metrics)

    # Sheet 2 — Simulated Inputs
    ws_inputs = wb.create_sheet("Simulated Inputs")
    _write_sheet(ws_inputs, inputs)

    # Sheet 3 — Assumptions (flatten to key/value pairs)
    ws_assump = wb.create_sheet("Assumptions")
    ws_assump.append(["Key", "Value"])
    for cell in ws_assump[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws_assump.column_dimensions["A"].width = 45
    ws_assump.column_dimensions["B"].width = 50

    def _flatten(obj: Any, prefix: str = "") -> list[tuple[str, str]]:
        pairs: list[tuple[str, str]] = []
        if isinstance(obj, dict):
            for k, v in obj.items():
                pairs.extend(_flatten(v, f"{prefix}.{k}" if prefix else k))
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                pairs.extend(_flatten(item, f"{prefix}[{i}]"))
        else:
            pairs.append((prefix, str(obj)))
        return pairs

    for key, value in _flatten(assumptions):
        ws_assump.append([key, value])

    out_path = out_dir / f"export_{date.today().isoformat()}.xlsx"
    out_dir.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Export financial outputs to JSON and Excel.")
    parser.add_argument("--metrics", type=Path, default=DEFAULT_METRICS)
    parser.add_argument("--inputs", type=Path, default=DEFAULT_INPUTS_CSV)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument("--assumptions", type=Path, default=ASSUMPTIONS_PATH)
    parser.add_argument(
        "--skip-excel",
        action="store_true",
        help="Skip Excel export (useful if openpyxl is not installed)",
    )
    args = parser.parse_args()

    assumptions = load_assumptions(args.assumptions)
    metrics = read_csv(args.metrics)
    inputs = read_csv(args.inputs)

    # ── JSON summary ──────────────────────────────────────────────────────────
    summary = {
        "generated_on": date.today().isoformat(),
        "assumptions_version": assumptions.get("model_meta", {}).get("version", "unknown"),
        "metrics": metrics,
    }
    json_path = args.out_dir / "financial_summary.json"
    write_json(summary, json_path)
    print(f"✓ Wrote JSON summary to {json_path}")

    # ── Excel export ──────────────────────────────────────────────────────────
    if not args.skip_excel:
        try:
            xlsx_path = write_excel(metrics, inputs, assumptions, args.out_dir)
            print(f"✓ Wrote Excel workbook to {xlsx_path}")
        except RuntimeError as exc:
            print(f"⚠  Excel export skipped: {exc}")
    else:
        print("⚠  Excel export skipped (--skip-excel flag set).")


if __name__ == "__main__":
    main()
